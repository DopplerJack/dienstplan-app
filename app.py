import streamlit as st
import pandas as pd
from ortools.sat.python import cp_model
import io
import datetime
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Dialyse Dienstplan", layout="wide")
st.title("🏥 Dienstplan-Generator: Dialyse (Stunden-Optimiert)")
st.markdown("Laden Sie Ihre Excel-Datei hoch. Die KI erstellt den Plan, achtet strikt auf die +/- 25h Grenze und verteilt die Stunden fair.")

uploaded_file = st.file_uploader("Excel-Tabelle hochladen (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    try:
        df_haupt = pd.read_excel(uploaded_file, sheet_name=0)
        df_haupt = df_haupt.astype(object) 
        df_haupt.fillna("Leer", inplace=True)
        
        try:
            df_regeln = pd.read_excel(uploaded_file, sheet_name=1)
            df_regeln = df_regeln.astype(object)
            df_regeln.fillna("", inplace=True)
        except:
            st.error("Fehler: Konnte das zweite Tabellenblatt 'Sonderregeln' nicht finden. Bitte legen Sie es an.")
            st.stop()
            
        st.success("Tabelle und Sonderregeln erfolgreich eingelesen!")
        
        if st.button("Dienstplan jetzt berechnen"):
            with st.spinner("Die KI balanciert nun die Stundenkonten aus... Dies kann einen Moment dauern."):
                
                model = cp_model.CpModel()
                
                mitarbeiter_liste = df_haupt['Name'].tolist()
                tage = [col for col in df_haupt.columns if col not in ['Name', 'Stundenausmaß', 'Berechnung Soll-Arbeitszeit', 'Übertrag Vormonat']]
                num_tage = len(tage)
                schichten = ['D1', 'V1', 'SL', 'D7', 'Frei', 'U', 'ZB', 'ÜZA', 'BA', 'FB']
                
                dienst_vars = {}
                for m in mitarbeiter_liste:
                    for t_idx in range(num_tage):
                        for s in schichten:
                            dienst_vars[(m, t_idx, s)] = model.NewBoolVar(f"{m}_{t_idx}_{s}")
                
                def parse_num(val):
                    try:
                        if pd.isna(val) or str(val).strip() == "Leer" or str(val).strip() == "": return 0.0
                        return float(str(val).replace(',', '.').replace('+', '').strip())
                    except:
                        return 0.0

                ma_daten = {}
                for index, row in df_haupt.iterrows():
                    m = row['Name']
                    ma_daten[m] = {
                        'ausmass_int': int(parse_num(row.get('Stundenausmaß', 0)) * 100),
                        'soll_int': int(parse_num(row.get('Berechnung Soll-Arbeitszeit', 0)) * 100),
                        'uebertrag_int': int(parse_num(row.get('Übertrag Vormonat', 0)) * 100)
                    }

                wochentage_idx = []
                samstage_idx = []
                for t_idx, tag_val in enumerate(tage):
                    if isinstance(tag_val, datetime.datetime):
                        wt = tag_val.weekday()
                    else:
                        try:
                            dt = pd.to_datetime(str(tag_val), dayfirst=True)
                            wt = dt.weekday()
                        except:
                            wt = t_idx % 7 
                            
                    wochentage_idx.append(wt)
                    if wt == 5: 
                        samstage_idx.append(t_idx)

                # --- 2. GRUNDREGELN & EXCEL-INPUT ---
                feste_eintraege = {} 
                
                for m in mitarbeiter_liste:
                    for t_idx in range(num_tage):
                        model.AddExactlyOne([dienst_vars[(m, t_idx, s)] for s in schichten])
                        
                for index, row in df_haupt.iterrows():
                    m = row['Name']
                    for t_idx, tag in enumerate(tage):
                        if wochentage_idx[t_idx] == 6: 
                            continue
                            
                        eintrag = str(row[tag]).strip()
                        if eintrag != "Leer" and eintrag != "":
                            feste_eintraege[(m, t_idx)] = eintrag 
                            if eintrag == "F": 
                                model.Add(dienst_vars[(m, t_idx, 'Frei')] == 1)
                            elif eintrag in schichten:
                                model.Add(dienst_vars[(m, t_idx, eintrag)] == 1)

                # ANPASSUNG: Die KI darf diese Kürzel niemals selbst vergeben!
                manuelle_kuerzel = ['SL', 'D7', 'U', 'ZB', 'ÜZA', 'BA', 'FB']
                for m in mitarbeiter_liste:
                    for t_idx in range(num_tage):
                        for kuerzel in manuelle_kuerzel:
                            # Wenn das Kürzel nicht explizit im Wunschplan stand, ist es für die KI verboten (0)
                            if not ((m, t_idx) in feste_eintraege and feste_eintraege[(m, t_idx)] == kuerzel):
                                model.Add(dienst_vars[(m, t_idx, kuerzel)] == 0)

                # --- 3. DAS ÜBERDRUCK-VENTIL ---
                straf_variablen = []
                fehlende_D1_vars = {}
                fehlende_V1_vars = {}
                
                for t_idx in range(num_tage):
                    wt = wochentage_idx[t_idx]
                    if wt == 6: 
                        for m in mitarbeiter_liste:
                            model.Add(dienst_vars[(m, t_idx, 'Frei')] == 1)
                        continue
                        
                    fehlend_d1 = model.NewIntVar(0, 7, f'fehlend_D1_{t_idx}')
                    fehlende_D1_vars[t_idx] = fehlend_d1
                    straf_variablen.append(fehlend_d1 * 1000000)
                    
                    if wt in [0, 1]: 
                        model.Add(sum(dienst_vars[(m, t_idx, 'D1')] for m in mitarbeiter_liste) + fehlend_d1 == 7)
                        
                        fehlend_v1 = model.NewIntVar(0, 1, f'fehlend_V1_{t_idx}')
                        fehlende_V1_vars[t_idx] = fehlend_v1
                        straf_variablen.append(fehlend_v1 * 1000000)
                        model.Add(sum(dienst_vars[(m, t_idx, 'V1')] for m in mitarbeiter_liste) + fehlend_v1 == 1)
                    else: 
                        model.Add(sum(dienst_vars[(m, t_idx, 'D1')] for m in mitarbeiter_liste) + fehlend_d1 == 7)
                        model.Add(sum(dienst_vars[(m, t_idx, 'V1')] for m in mitarbeiter_liste) == 0)

                # --- 4. FIXE GLOBALE REGELN ---
                mondays = [t for t, wt in enumerate(wochentage_idx) if wt == 0]
                
                for m in mitarbeiter_liste:
                    for t in range(num_tage - 3):
                        model.Add(sum(dienst_vars[(m, t+i, 'D1')] for i in range(4)) <= 3)
                        
                    for t in range(num_tage - 6):
                        model.Add(sum(dienst_vars[(m, t+i, s)] for i in range(7) for s in ['D1', 'V1', 'SL', 'D7']) <= 4)
                        
                    model.Add(sum(dienst_vars[(m, t, 'V1')] for t in range(num_tage)) <= 1)
                        
                    if len(samstage_idx) > 0:
                        model.Add(sum(dienst_vars[(m, t, 'D1')] for t in samstage_idx) <= 3) 
                        dritter_sat = model.NewBoolVar(f"DritterSat_{m}")
                        model.Add(sum(dienst_vars[(m, t, 'D1')] for t in samstage_idx) <= 2 + dritter_sat)
                        straf_variablen.append(dritter_sat * 50000) 

                    for mo in mondays:
                        if mo + 6 < num_tage: 
                            week_pairs = []
                            for i in range(6):
                                t1 = mo + i
                                t2 = mo + i + 1
                                
                                work1 = model.NewBoolVar(f"work_{m}_{t1}")
                                model.Add(work1 == sum(dienst_vars[(m, t1, s)] for s in ['D1', 'V1', 'SL', 'D7']))
                                
                                work2 = model.NewBoolVar(f"work_{m}_{t2}")
                                model.Add(work2 == sum(dienst_vars[(m, t2, s)] for s in ['D1', 'V1', 'SL', 'D7']))
                                
                                pair_off = model.NewBoolVar(f"pairoff_{m}_{t1}")
                                model.Add(work1 + work2 == 0).OnlyEnforceIf(pair_off)
                                model.Add(work1 + work2 > 0).OnlyEnforceIf(pair_off.Not())
                                
                                week_pairs.append(pair_off)
                            
                            model.AddBoolOr(week_pairs)

                # --- 5. INDIVIDUELLE REGELN ---
                wt_map = {"Montag": 0, "Dienstag": 1, "Mittwoch": 2, "Donnerstag": 3, "Freitag": 4, "Samstag": 5, "Sonntag": 6}
                
                for index, row in df_regeln.iterrows():
                    m = row['Name']
                    if m not in mitarbeiter_liste: continue
                    
                    freier_tag_str = str(row.get('Fester freier Tag', '')).strip()
                    if freier_tag_str:
                        feste_freie_tage = [t.strip() for t in freier_tag_str.split(',')]
                        for tag_name in feste_freie_tage:
                            if tag_name in wt_map:
                                ziel_wt = wt_map[tag_name]
                                for t_idx in range(num_tage):
                                    if wochentage_idx[t_idx] == ziel_wt:
                                        if (m, t_idx) not in feste_eintraege:
                                            model.Add(dienst_vars[(m, t_idx, 'Frei')] == 1)

                    fester_dienst_str = str(row.get('Fester Tag Dienst', '')).strip()
                    if fester_dienst_str:
                        feste_dienst_tage = [t.strip() for t in fester_dienst_str.split(',')]
                        for tag_name in feste_dienst_tage:
                            if tag_name in wt_map:
                                ziel_wt = wt_map[tag_name]
                                for t_idx in range(num_tage):
                                    if wochentage_idx[t_idx] == ziel_wt:
                                        if (m, t_idx) not in feste_eintraege:
                                            model.Add(dienst_vars[(m, t_idx, 'D1')] == 1)
                                
                    if str(row.get('Keine V1', '')).strip().lower() == 'ja':
                        for t_idx in range(num_tage):
                            model.Add(dienst_vars[(m, t_idx, 'V1')] == 0)

                    if str(row.get('Immer ein V1-Dienst', '')).strip().lower() == 'ja':
                        model.Add(sum(dienst_vars[(m, t_idx, 'V1')] for t_idx in range(num_tage)) == 1)
                            
                    if str(row.get('Max 1 D1 am Stück', '')).strip().lower() == 'ja':
                        for t_idx in range(num_tage - 1):
                            model.Add(dienst_vars[(m, t_idx, 'D1')] + dienst_vars[(m, t_idx+1, 'D1')] <= 1)

                    if str(row.get('Freitag frei vor Samstag-D1', '')).strip().lower() == 'ja':
                        for t_idx in samstage_idx:
                            if t_idx > 0: 
                                model.AddImplication(dienst_vars[(m, t_idx, 'D1')], dienst_vars[(m, t_idx-1, 'Frei')])

                    if str(row.get('Keine Samstag/Montag Konstellation', '')).strip().lower() == 'ja':
                        for t_idx in samstage_idx:
                            if t_idx + 2 < num_tage: 
                                model.AddImplication(dienst_vars[(m, t_idx, 'D1')], dienst_vars[(m, t_idx+2, 'Frei')])

                    if str(row.get('Bevorzuge 2er D1-Blöcke', '')).strip().lower() == 'ja':
                        for t_idx in range(num_tage):
                            d_heute = dienst_vars[(m, t_idx, 'D1')]
                            d_gestern = dienst_vars[(m, t_idx-1, 'D1')] if t_idx > 0 else 0
                            d_morgen = dienst_vars[(m, t_idx+1, 'D1')] if t_idx < num_tage - 1 else 0
                            
                            val_iso = model.NewIntVar(-2, 1, f"iso_{m}_{t_idx}")
                            model.Add(val_iso == d_heute - d_gestern - d_morgen)
                            is_iso = model.NewBoolVar(f"is_iso_bool_{m}_{t_idx}")
                            model.Add(val_iso == 1).OnlyEnforceIf(is_iso)
                            model.Add(val_iso != 1).OnlyEnforceIf(is_iso.Not())
                            straf_variablen.append(is_iso * 15000)
                            
                            val_3 = model.NewIntVar(0, 3, f"3blk_{m}_{t_idx}")
                            model.Add(val_3 == d_gestern + d_heute + d_morgen)
                            is_3blk = model.NewBoolVar(f"is_3blk_bool_{m}_{t_idx}")
                            model.Add(val_3 == 3).OnlyEnforceIf(is_3blk)
                            model.Add(val_3 != 3).OnlyEnforceIf(is_3blk.Not())
                            straf_variablen.append(is_3blk * 10000)

                    if str(row.get('Bevorzuge 3er D1-Blöcke', '')).strip().lower() == 'ja':
                        for t_idx in range(num_tage):
                            d_heute = dienst_vars[(m, t_idx, 'D1')]
                            d_gestern = dienst_vars[(m, t_idx-1, 'D1')] if t_idx > 0 else 0
                            d_morgen = dienst_vars[(m, t_idx+1, 'D1')] if t_idx < num_tage - 1 else 0
                            
                            val_iso = model.NewIntVar(-2, 1, f"iso3_{m}_{t_idx}")
                            model.Add(val_iso == d_heute - d_gestern - d_morgen)
                            is_iso = model.NewBoolVar(f"is_iso3_bool_{m}_{t_idx}")
                            model.Add(val_iso == 1).OnlyEnforceIf(is_iso)
                            model.Add(val_iso != 1).OnlyEnforceIf(is_iso.Not())
                            straf_variablen.append(is_iso * 15000)
                            
                        for t_idx in range(num_tage - 1):
                            d_heute = dienst_vars[(m, t_idx, 'D1')]
                            d_morgen = dienst_vars[(m, t_idx+1, 'D1')]
                            d_gestern = dienst_vars[(m, t_idx-1, 'D1')] if t_idx > 0 else 0
                            d_uebermorgen = dienst_vars[(m, t_idx+2, 'D1')] if t_idx < num_tage - 2 else 0
                            
                            val_2blk = model.NewIntVar(-2, 2, f"2blk_{m}_{t_idx}")
                            model.Add(val_2blk == d_heute + d_morgen - d_gestern - d_uebermorgen)
                            is_2blk = model.NewBoolVar(f"is_2blk_bool_{m}_{t_idx}")
                            model.Add(val_2blk == 2).OnlyEnforceIf(is_2blk)
                            model.Add(val_2blk != 2).OnlyEnforceIf(is_2blk.Not())
                            straf_variablen.append(is_2blk * 10000)

                # --- 6. WEICHE REGELN & OPTIMIERUNG ---
                for m in mitarbeiter_liste:
                    for i in range(len(samstage_idx) - 1):
                        sat1 = samstage_idx[i]
                        sat2 = samstage_idx[i+1]
                        doppel_sat = model.NewBoolVar(f"DoppelSat_{m}_{sat1}")
                        model.Add(dienst_vars[(m, sat1, 'D1')] + dienst_vars[(m, sat2, 'D1')] == 2).OnlyEnforceIf(doppel_sat)
                        straf_variablen.append(doppel_sat * 10000) 

                # --- 7. STUNDENKONTO & FAIRNESS ---
                for m in mitarbeiter_liste:
                    ausmass_int = ma_daten[m]['ausmass_int']
                    soll_int = ma_daten[m]['soll_int']
                    uebertrag_int = ma_daten[m]['uebertrag_int']
                    tageswert_int = int(ausmass_int / 5.0) if ausmass_int > 0 else 0
                    
                    schicht_dauer_liste = []
                    for t_idx in range(num_tage):
                        schicht_dauer_liste.append(dienst_vars[(m, t_idx, 'D1')] * 1125)
                        schicht_dauer_liste.append(dienst_vars[(m, t_idx, 'V1')] * 700)
                        schicht_dauer_liste.append(dienst_vars[(m, t_idx, 'SL')] * 900)
                        schicht_dauer_liste.append(dienst_vars[(m, t_idx, 'D7')] * 825)
                        for abw in ['U', 'ZB', 'ÜZA', 'BA', 'FB']:
                            schicht_dauer_liste.append(dienst_vars[(m, t_idx, abw)] * tageswert_int)
                            
                    plan_std_var = model.NewIntVar(0, 50000, f'plan_std_{m}')
                    model.Add(plan_std_var == sum(schicht_dauer_liste))
                    
                    if soll_int > 0: 
                        model.Add(plan_std_var >= soll_int - 2500)
                        model.Add(plan_std_var <= soll_int + 2500)
                    
                    diff_var = model.NewIntVar(-50000, 50000, f'diff_{m}')
                    model.Add(diff_var == plan_std_var + uebertrag_int - soll_int)
                    
                    abs_diff = model.NewIntVar(0, 50000, f'abs_diff_{m}')
                    model.AddAbsEquality(abs_diff, diff_var)
                    
                    straf_variablen.append(abs_diff * 10)

                model.Minimize(sum(straf_variablen))

                solver = cp_model.CpSolver()
                solver.parameters.max_time_in_seconds = 90.0 
                status = solver.Solve(model)

                if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
                    
                    warnungen = []
                    for t_idx in fehlende_D1_vars:
                        fehlt_d1 = solver.Value(fehlende_D1_vars[t_idx])
                        if fehlt_d1 > 0:
                            datum = tage[t_idx]
                            if isinstance(datum, datetime.datetime):
                                datum = datum.strftime('%d.%m.%Y')
                            warnungen.append(f"⚠️ **{datum}**: Es fehlen **{fehlt_d1}** Mitarbeiter für den D1-Dienst.")
                            
                    for t_idx in fehlende_V1_vars:
                        fehlt_v1 = solver.Value(fehlende_V1_vars[t_idx])
                        if fehlt_v1 > 0:
                            datum = tage[t_idx]
                            if isinstance(datum, datetime.datetime):
                                datum = datum.strftime('%d.%m.%Y')
                            warnungen.append(f"⚠️ **{datum}**: Der V1-Dienst konnte nicht besetzt werden.")

                    if len(warnungen) > 0:
                        st.warning("Der Dienstplan wurde erstellt, aber es gibt personelle Engpässe! Die Station ist an folgenden Tagen unterbesetzt:")
                        for w in warnungen:
                            st.write(w)
                    else:
                        st.success("🎉 Plan erfolgreich berechnet! Die Stunden sind gerecht verteilt.")

                    ausgabe_df = df_haupt.copy()

                    plan_stunden_liste = []
                    abweichung_liste = []

                    for index, row in df_haupt.iterrows():
                        m = row['Name']
                        stundenausmass = parse_num(row.get('Stundenausmaß', 0))
                        soll_arbeitszeit = parse_num(row.get('Berechnung Soll-Arbeitszeit', 0))
                        tageswert_abwesenheit = stundenausmass / 5.0
                        plan_std_aktuell = 0.0
                        
                        for t_idx, tag in enumerate(tage):
                            for s in schichten:
                                if solver.Value(dienst_vars[(m, t_idx, s)]) == 1:
                                    
                                    if s == 'Frei':
                                        if (m, t_idx) in feste_eintraege and feste_eintraege[(m, t_idx)] == 'F':
                                            final_eintrag = 'F'
                                        else:
                                            final_eintrag = '--'
                                    else:
                                        final_eintrag = s
                                        
                                    ausgabe_df.at[index, tag] = final_eintrag
                                    
                                    if s == 'D1': plan_std_aktuell += 11.25
                                    elif s == 'V1': plan_std_aktuell += 7.0
                                    elif s == 'SL': plan_std_aktuell += 9.0
                                    elif s == 'D7': plan_std_aktuell += 8.25
                                    elif s in ['U', 'ZB', 'ÜZA', 'BA', 'FB']: plan_std_aktuell += tageswert_abwesenheit
                        
                        plan_stunden_liste.append(round(plan_std_aktuell, 2))
                        abweichung_liste.append(round(plan_std_aktuell - soll_arbeitszeit, 2))

                    if 'Übertrag Vormonat' in ausgabe_df.columns:
                        col_idx = ausgabe_df.columns.get_loc('Übertrag Vormonat') + 1
                    else:
                        col_idx = 3 
                        
                    ausgabe_df.insert(col_idx, 'Plan-Stunden', plan_stunden_liste)
                    ausgabe_df.insert(col_idx + 1, 'Abweichung vom Soll', abweichung_liste)

                    st.dataframe(ausgabe_df.head(10))
                    
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        ausgabe_df.to_excel(writer, index=False, sheet_name="Dienstplan")
                        
                        workbook = writer.book
                        worksheet = writer.sheets['Dienstplan']
                        
                        fill_d1 = PatternFill(start_color="40E0D0", end_color="40E0D0", fill_type="solid")
                        fill_v1 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        fill_gruen = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        fill_rot = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
                        fill_grau = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        
                        for row in worksheet.iter_rows(min_row=2): 
                            for cell in row:
                                val = str(cell.value).strip() if cell.value else ""
                                if val == 'D1':
                                    cell.fill = fill_d1
                                elif val == 'V1':
                                    cell.fill = fill_v1
                                elif val in ['U', 'ZB', 'ÜZA', 'BA', 'FB']:
                                    cell.fill = fill_gruen
                                elif val == 'F':
                                    cell.fill = fill_rot
                                elif val == '--':
                                    cell.fill = fill_grau

                    st.download_button(
                        label="📥 Fertigen Dienstplan (inkl. Farben & Stunden) herunterladen", 
                        data=buffer.getvalue(), 
                        file_name="Dienstplan_Fertig_Farbig.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error("🚨 Kritischer Fehler! Die Mathematik widerspricht sich komplett.")
    except Exception as e:
        st.error(f"Ein Fehler ist aufgetreten: {e}")
